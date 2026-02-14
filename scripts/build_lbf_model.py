#!/usr/bin/env python3
"""Build LBF Financial Model Excel Workbook"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(bold=True, size=11, color="2F5496")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
money_fmt = '#,##0'
pct_fmt = '0%'
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
neg_font = Font(bold=True, color="FF0000")
pos_font = Font(bold=True, color="008000")
thin_border = Border(
    bottom=Side(style='thin', color='CCCCCC')
)

months = ['Mar 26', 'Apr 26', 'May 26', 'Jun 26', 'Jul 26', 'Aug 26',
          'Sep 26', 'Oct 26', 'Nov 26', 'Dec 26', 'Jan 27', 'Feb 27']

# ═══════════════════════════════════════════
# SHEET 1: P&L Summary
# ═══════════════════════════════════════════
ws = wb.active
ws.title = "P&L Summary"
ws.sheet_properties.tabColor = "2F5496"

# Title
ws.merge_cells('A1:N1')
ws['A1'] = 'LBF ENTERPRISE — 12-MONTH P&L PROJECTION'
ws['A1'].font = Font(bold=True, size=16, color="2F5496")

ws.merge_cells('A2:N2')
ws['A2'] = 'Helios Architect LBF | Prepared 2026-02-13'
ws['A2'].font = Font(size=10, color="808080")

# Column widths
ws.column_dimensions['A'].width = 28
for i in range(2, 15):
    ws.column_dimensions[get_column_letter(i)].width = 12

# Headers
row = 4
ws.cell(row=row, column=1, value='').fill = header_fill
for i, m in enumerate(months):
    c = ws.cell(row=row, column=i+2, value=m)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
c = ws.cell(row=row, column=14, value='YEAR 1')
c.font = header_font
c.fill = header_fill
c.alignment = Alignment(horizontal='center')

# Revenue Section
row = 5
ws.cell(row=row, column=1, value='REVENUE').font = section_font
for i in range(1, 15):
    ws.cell(row=row, column=i).fill = section_fill

revenue_data = {
    'WEMS MCP Server':     [80, 150, 290, 450, 650, 900, 1100, 1350, 1600, 1850, 2100, 2399],
    'ClawHub Marketplace':  [180, 350, 540, 750, 950, 1100, 1400, 1700, 2000, 2300, 2550, 2800],
    'Brain-DB SaaS':       [106, 300, 585, 800, 1050, 1268, 1600, 1900, 2200, 2500, 2750, 2980],
    'LLM Fleet Mgmt':      [99, 250, 596, 900, 1600, 2391, 2800, 3200, 3500, 3800, 4150, 4483],
    'AUGUR Signals':       [296, 700, 1236, 1500, 1800, 2078, 2400, 2700, 2900, 3100, 3350, 3612],
    'Consulting':          [1000, 3000, 5000, 6000, 7000, 8000, 8500, 9000, 9500, 10000, 11000, 12000],
}

for name, values in revenue_data.items():
    row += 1
    ws.cell(row=row, column=1, value=f'  {name}').border = thin_border
    annual = 0
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i+2, value=v)
        c.number_format = money_fmt
        c.border = thin_border
        annual += v
    c = ws.cell(row=row, column=14, value=annual)
    c.number_format = money_fmt
    c.border = thin_border

# Total Revenue
row += 1
ws.cell(row=row, column=1, value='TOTAL REVENUE').font = total_font
for i in range(1, 15):
    ws.cell(row=row, column=i).fill = total_fill
total_rev_row = row
for i in range(12):
    total = sum(v[i] for v in revenue_data.values())
    c = ws.cell(row=row, column=i+2, value=total)
    c.number_format = money_fmt
    c.font = total_font
c = ws.cell(row=row, column=14, value=sum(sum(v) for v in revenue_data.values()))
c.number_format = money_fmt
c.font = total_font

# Expenses Section
row += 2
ws.cell(row=row, column=1, value='EXPENSES').font = section_font
for i in range(1, 15):
    ws.cell(row=row, column=i).fill = section_fill

expense_data = {
    'Anthropic/OpenAI API': [2400, 2400, 2400, 2500, 2600, 2800, 2900, 3000, 3000, 3100, 3100, 3200],
    'Infrastructure':       [282, 282, 300, 320, 350, 400, 420, 440, 460, 480, 490, 500],
    'SaaS Subscriptions':   [142, 142, 160, 170, 180, 200, 210, 220, 230, 240, 245, 250],
    'Payment Processing':   [51, 138, 239, 302, 378, 456, 516, 576, 629, 683, 751, 820],
    'Utilities':            [150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150],
    'Marketing/CAC':        [100, 150, 300, 350, 400, 500, 550, 600, 650, 700, 750, 800],
    'Legal/Compliance':     [0, 0, 200, 200, 200, 200, 200, 200, 200, 200, 200, 200],
}

for name, values in expense_data.items():
    row += 1
    ws.cell(row=row, column=1, value=f'  {name}').border = thin_border
    annual = 0
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i+2, value=v)
        c.number_format = money_fmt
        c.border = thin_border
        annual += v
    c = ws.cell(row=row, column=14, value=annual)
    c.number_format = money_fmt
    c.border = thin_border

# Total Expenses
row += 1
ws.cell(row=row, column=1, value='TOTAL EXPENSES').font = total_font
for i in range(1, 15):
    ws.cell(row=row, column=i).fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
total_exp_row = row
for i in range(12):
    total = sum(v[i] for v in expense_data.values())
    c = ws.cell(row=row, column=i+2, value=total)
    c.number_format = money_fmt
    c.font = total_font
c = ws.cell(row=row, column=14, value=sum(sum(v) for v in expense_data.values()))
c.number_format = money_fmt
c.font = total_font

# Net Income
row += 2
ws.cell(row=row, column=1, value='NET INCOME').font = Font(bold=True, size=13, color="2F5496")
for i in range(1, 15):
    ws.cell(row=row, column=i).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
net_row = row
cumulative = 0
annual_net = 0
for i in range(12):
    rev = sum(v[i] for v in revenue_data.values())
    exp = sum(v[i] for v in expense_data.values())
    net = rev - exp
    annual_net += net
    c = ws.cell(row=row, column=i+2, value=net)
    c.number_format = money_fmt
    c.font = pos_font if net >= 0 else neg_font
c = ws.cell(row=row, column=14, value=annual_net)
c.number_format = money_fmt
c.font = pos_font

# Cumulative
row += 1
ws.cell(row=row, column=1, value='CUMULATIVE').font = Font(bold=True, size=11)
cumulative = 0
for i in range(12):
    rev = sum(v[i] for v in revenue_data.values())
    exp = sum(v[i] for v in expense_data.values())
    cumulative += (rev - exp)
    c = ws.cell(row=row, column=i+2, value=cumulative)
    c.number_format = money_fmt
    c.font = pos_font if cumulative >= 0 else neg_font

# Margin %
row += 1
ws.cell(row=row, column=1, value='NET MARGIN %').font = Font(bold=True, size=11)
for i in range(12):
    rev = sum(v[i] for v in revenue_data.values())
    exp = sum(v[i] for v in expense_data.values())
    margin = (rev - exp) / rev if rev > 0 else 0
    c = ws.cell(row=row, column=i+2, value=margin)
    c.number_format = '0%'

# ═══════════════════════════════════════════
# SHEET 2: Unit Economics
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Unit Economics")
ws2.sheet_properties.tabColor = "548235"
ws2.column_dimensions['A'].width = 24
for i in range(2, 9):
    ws2.column_dimensions[get_column_letter(i)].width = 16

ws2.merge_cells('A1:H1')
ws2['A1'] = 'UNIT ECONOMICS BY PRODUCT'
ws2['A1'].font = Font(bold=True, size=14, color="548235")

headers = ['Product', 'Price Range', 'COGS/User', 'Gross Margin', 'Est. CAC', 'LTV (12mo)', 'LTV:CAC', 'Monthly Churn']
for i, h in enumerate(headers):
    c = ws2.cell(row=3, column=i+1, value=h)
    c.font = header_font
    c.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    c.alignment = Alignment(horizontal='center')

unit_data = [
    ['WEMS MCP Server', '$0-30/mo', '$0.50', 0.95, '$10', '$184', '18:1', 0.08],
    ['ClawHub Marketplace', '$5-29/mo', '$0.10', 0.90, '$8', '$230', '29:1', 0.10],
    ['Brain-DB SaaS', '$19-199/mo', '$5.00', 0.87, '$35', '$529', '15:1', 0.05],
    ['LLM Fleet Mgmt', '$99-999/mo', '$30.00', 0.85, '$50', '$2,149', '43:1', 0.05],
    ['AUGUR Signals', '$49-299/mo', '$10.00', 0.85, '$50', '$490', '10:1', 0.20],
    ['Consulting', '$150/hr', '$0', 0.98, '$0', 'N/A', 'N/A', 'N/A'],
]

for r, data in enumerate(unit_data):
    for c_idx, val in enumerate(data):
        cell = ws2.cell(row=r+4, column=c_idx+1, value=val)
        if isinstance(val, float) and val < 1:
            cell.number_format = '0%'
        cell.border = thin_border
        if c_idx > 0:
            cell.alignment = Alignment(horizontal='center')

# ═══════════════════════════════════════════
# SHEET 3: Cost Structure  
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Cost Structure")
ws3.sheet_properties.tabColor = "BF4B28"
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 40

ws3.merge_cells('A1:E1')
ws3['A1'] = 'CURRENT COST STRUCTURE (Monthly)'
ws3['A1'].font = Font(bold=True, size=14, color="BF4B28")

headers = ['Category', 'Monthly Cost', 'Annual Cost', '% of Total', 'Notes']
for i, h in enumerate(headers):
    c = ws3.cell(row=3, column=i+1, value=h)
    c.font = header_font
    c.fill = PatternFill(start_color="BF4B28", end_color="BF4B28", fill_type="solid")

costs = [
    ['Anthropic Claude API', 2100, 25200, 0.80, 'Opus 4.6 primary (~$70/day avg)'],
    ['OpenAI API', 300, 3600, 0.11, 'Supplementary usage'],
    ['Cloud/VPS Hosting', 40, 480, 0.02, 'Light cloud footprint'],
    ['Domain Registrations', 12, 144, 0.005, 'Multiple domains'],
    ['GitHub/SaaS', 30, 360, 0.01, 'Dev tools, subscriptions'],
    ['Internet Service', 100, 1200, 0.04, '85% business use = $1,020 deductible'],
    ['Electricity (servers)', 50, 600, 0.02, '4-node homelab fleet'],
]

for r, data in enumerate(costs):
    for c_idx, val in enumerate(data):
        cell = ws3.cell(row=r+4, column=c_idx+1, value=val)
        if c_idx in [1, 2]:
            cell.number_format = '$#,##0'
        elif c_idx == 3:
            cell.number_format = '0%'
        cell.border = thin_border

# Total
total_row = len(costs) + 4
ws3.cell(row=total_row, column=1, value='TOTAL').font = total_font
ws3.cell(row=total_row, column=2, value=2632).font = total_font
ws3.cell(row=total_row, column=2).number_format = '$#,##0'
ws3.cell(row=total_row, column=3, value=31584).font = total_font
ws3.cell(row=total_row, column=3).number_format = '$#,##0'
ws3.cell(row=total_row, column=4, value=1.0)
ws3.cell(row=total_row, column=4).number_format = '0%'
for i in range(1, 6):
    ws3.cell(row=total_row, column=i).fill = total_fill

# Cost optimization section
opt_row = total_row + 3
ws3.merge_cells(f'A{opt_row}:E{opt_row}')
ws3.cell(row=opt_row, column=1, value='COST OPTIMIZATION LEVERS').font = Font(bold=True, size=12, color="BF4B28")

headers2 = ['Lever', 'Monthly Savings', 'Annual Savings', 'Effort', 'Timeline']
for i, h in enumerate(headers2):
    c = ws3.cell(row=opt_row+1, column=i+1, value=h)
    c.font = header_font
    c.fill = PatternFill(start_color="BF4B28", end_color="BF4B28", fill_type="solid")

optimizations = [
    ['Local LLM routing (Ollama)', 800, 9600, 'Medium', '2-4 weeks'],
    ['Prompt caching', 300, 3600, 'Low', '1 week'],
    ['Model downgrade (Sonnet for non-critical)', 500, 6000, 'Low', 'Immediate'],
    ['RTX 5090 (local inference)', 1200, 14400, 'High ($2K+ CapEx)', '4-8 weeks'],
]

for r, data in enumerate(optimizations):
    for c_idx, val in enumerate(data):
        cell = ws3.cell(row=opt_row+2+r, column=c_idx+1, value=val)
        if c_idx in [1, 2]:
            cell.number_format = '$#,##0'
        cell.border = thin_border

# ═══════════════════════════════════════════
# SHEET 4: Scenario Analysis
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Scenarios")
ws4.sheet_properties.tabColor = "7030A0"
ws4.column_dimensions['A'].width = 20
for i in range(2, 7):
    ws4.column_dimensions[get_column_letter(i)].width = 18

ws4.merge_cells('A1:F1')
ws4['A1'] = 'SCENARIO ANALYSIS'
ws4['A1'].font = Font(bold=True, size=14, color="7030A0")

headers = ['Scenario', 'Revenue Multiple', 'Year 1 Revenue', 'Year 1 Expenses', 'Year 1 Net', 'Monthly at M12']
for i, h in enumerate(headers):
    c = ws4.cell(row=3, column=i+1, value=h)
    c.font = header_font
    c.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    c.alignment = Alignment(horizontal='center')

scenarios = [
    ['Pessimistic', 0.40, 68000, 45000, 23000, 11310],
    ['Conservative', 0.60, 102000, 48000, 54000, 16964],
    ['Base Case', 1.00, 170000, 52000, 118000, 28274],
    ['Optimistic', 1.50, 255000, 58000, 197000, 42411],
    ['Bull Case', 2.00, 340000, 65000, 275000, 56548],
]

for r, data in enumerate(scenarios):
    for c_idx, val in enumerate(data):
        cell = ws4.cell(row=r+4, column=c_idx+1, value=val)
        if c_idx == 1:
            cell.number_format = '0%'
        elif c_idx >= 2:
            cell.number_format = '$#,##0'
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

# Key Assumptions
ws4.cell(row=11, column=1, value='KEY ASSUMPTIONS').font = Font(bold=True, size=12, color="7030A0")
assumptions = [
    ['Free-to-paid conversion', '5%', 'Industry avg for dev tools'],
    ['SaaS monthly churn', '5-8%', 'B2B SaaS typical'],
    ['Signal service churn', '15-25%', 'High for signal services'],
    ['Consulting utilization', '10-15 hrs/wk', 'Preserves product dev time'],
    ['API cost growth', '3%/quarter', 'Sub-linear with revenue'],
    ['MCP ecosystem growth', '50%+ YoY', 'Enterprise adoption accelerating'],
    ['Payment processing', '2.9%', 'Stripe standard rate'],
]

headers_a = ['Assumption', 'Value', 'Notes']
for i, h in enumerate(headers_a):
    c = ws4.cell(row=12, column=i+1, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

for r, data in enumerate(assumptions):
    for c_idx, val in enumerate(data):
        ws4.cell(row=r+13, column=c_idx+1, value=val).border = thin_border

# ═══════════════════════════════════════════
# SHEET 5: Revenue Mix
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("Revenue Mix")
ws5.sheet_properties.tabColor = "ED7D31"
ws5.column_dimensions['A'].width = 24
for i in range(2, 8):
    ws5.column_dimensions[get_column_letter(i)].width = 14

ws5.merge_cells('A1:G1')
ws5['A1'] = 'REVENUE MIX EVOLUTION'
ws5['A1'].font = Font(bold=True, size=14, color="ED7D31")

headers = ['Stream', 'M1 $', 'M1 %', 'M6 $', 'M6 %', 'M12 $', 'M12 %']
for i, h in enumerate(headers):
    c = ws5.cell(row=3, column=i+1, value=h)
    c.font = header_font
    c.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    c.alignment = Alignment(horizontal='center')

m1_total = 1761
m6_total = 15737
m12_total = 28274

mix_data = [
    ['WEMS MCP Server', 80, 80/m1_total, 900, 900/m6_total, 2399, 2399/m12_total],
    ['ClawHub Marketplace', 180, 180/m1_total, 1100, 1100/m6_total, 2800, 2800/m12_total],
    ['Brain-DB SaaS', 106, 106/m1_total, 1268, 1268/m6_total, 2980, 2980/m12_total],
    ['LLM Fleet Mgmt', 99, 99/m1_total, 2391, 2391/m6_total, 4483, 4483/m12_total],
    ['AUGUR Signals', 296, 296/m1_total, 2078, 2078/m6_total, 3612, 3612/m12_total],
    ['Consulting', 1000, 1000/m1_total, 8000, 8000/m6_total, 12000, 12000/m12_total],
]

for r, data in enumerate(mix_data):
    for c_idx, val in enumerate(data):
        cell = ws5.cell(row=r+4, column=c_idx+1, value=val)
        if c_idx in [2, 4, 6]:
            cell.number_format = '0%'
        elif c_idx in [1, 3, 5]:
            cell.number_format = '$#,##0'
        cell.border = thin_border
        if c_idx > 0:
            cell.alignment = Alignment(horizontal='center')

# Totals
total_r = len(mix_data) + 4
for i in range(1, 8):
    ws5.cell(row=total_r, column=i).fill = total_fill
ws5.cell(row=total_r, column=1, value='TOTAL').font = total_font
ws5.cell(row=total_r, column=2, value=m1_total).number_format = '$#,##0'
ws5.cell(row=total_r, column=3, value=1.0).number_format = '0%'
ws5.cell(row=total_r, column=4, value=m6_total).number_format = '$#,##0'
ws5.cell(row=total_r, column=5, value=1.0).number_format = '0%'
ws5.cell(row=total_r, column=6, value=m12_total).number_format = '$#,##0'
ws5.cell(row=total_r, column=7, value=1.0).number_format = '0%'

# Strategic note
note_r = total_r + 2
ws5.merge_cells(f'A{note_r}:G{note_r}')
ws5.cell(row=note_r, column=1, value='⚠️ STRATEGIC NOTE: Consulting drops from 57% to 42% of revenue M1→M12. Target: <25% by M18.').font = Font(italic=True, color="BF4B28")

# Save
outpath = '/home/bonsaihorn/.openclaw/workspace/LBF_Financial_Model.xlsx'
wb.save(outpath)
print(f"Saved: {outpath}")
print("Sheets: P&L Summary, Unit Economics, Cost Structure, Scenarios, Revenue Mix")
